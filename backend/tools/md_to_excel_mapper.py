from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string


def parse_md(path: Path) -> list[tuple[str, str]]:
    text = path.read_text(encoding="utf-8")
    lines = text.splitlines()

    entries: list[tuple[str, str]] = []

    first_heading_idx = None
    for i, line in enumerate(lines):
        if re.match(r"^#{1,6}\s+", line):
            first_heading_idx = i
            break

    if first_heading_idx is None:
        return entries

    front = "\n".join(lines[:first_heading_idx]).strip()
    if front:
        entries.append(("Document Header", front))

    i = first_heading_idx
    while i < len(lines):
        m = re.match(r"^(#{1,6})\s+(.*)$", lines[i])
        if not m:
            i += 1
            continue

        title = m.group(2).strip()
        j = i + 1
        body_lines: list[str] = []
        while j < len(lines) and not re.match(r"^#{1,6}\s+", lines[j]):
            body_lines.append(lines[j])
            j += 1

        body = "\n".join(body_lines).strip()
        entries.append((title, body))
        i = j

    return entries


def safe_set(ws, row: int, col: int, value: str | None) -> None:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if (row, col) in merged_range.cells:
                anchor = ws.cell(merged_range.min_row, merged_range.min_col)
                anchor.value = value
                return
        return
    cell.value = value


def write_export_sheet(
    template_path: Path,
    out_path: Path,
    entries: list[tuple[str, str]],
    title_col_letter: str,
    body_col_letter: str,
    type_col_letter: str = "B",
    type_value: str = "Information",
    start_row: int = 4,
    clear_until_row: int = 1200,
) -> None:
    wb = load_workbook(template_path)
    ws = wb["Export"]

    title_col = column_index_from_string(title_col_letter.upper())
    body_col = column_index_from_string(body_col_letter.upper())
    type_col = column_index_from_string(type_col_letter.upper())

    for row in range(start_row, clear_until_row):
        safe_set(ws, row, type_col, None)
        safe_set(ws, row, title_col, None)
        safe_set(ws, row, body_col, None)

    row = start_row
    for title, body in entries:
        if not title and not body:
            continue
        safe_set(ws, row, type_col, type_value)
        safe_set(ws, row, title_col, title)
        safe_set(ws, row, body_col, body)
        row += 1

    wb.save(out_path)


def filter_hsi(entries: list[tuple[str, str]]) -> list[tuple[str, str]]:
    return [
        (title, body)
        for title, body in entries
        if ("hardware-software interface" in title.lower() or "hsi" in title.lower())
    ]


def filter_verification(entries: list[tuple[str, str]]) -> list[tuple[str, str]]:
    return [(title, body) for title, body in entries if "verification" in title.lower()]


def write_preview_csv(
    out_path: Path,
    entries: list[tuple[str, str]],
    title_col_letter: str,
    body_col_letter: str,
    type_col_letter: str = "B",
    type_value: str = "Information",
    start_row: int = 4,
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "excel_row",
                "type_column",
                "type_value",
                "title_column",
                "title",
                "body_column",
                "body",
            ]
        )

        row = start_row
        for title, body in entries:
            if not title and not body:
                continue
            writer.writerow(
                [
                    row,
                    type_col_letter,
                    type_value,
                    title_col_letter,
                    title,
                    body_col_letter,
                    body,
                ]
            )
            row += 1


def main() -> None:
    parser = argparse.ArgumentParser(description="Map Markdown headings/body to Excel Export sheet columns")
    parser.add_argument("--base-dir", required=True, help="Directory containing markdown files and outputs")
    parser.add_argument("--req-template", required=True, help="Requirements Excel template path")
    parser.add_argument("--arch-template", required=True, help="Architecture Excel template path")

    parser.add_argument("--req-title-col", default="C")
    parser.add_argument("--req-body-col", default="M")
    parser.add_argument("--arch-title-col", default="C")
    parser.add_argument("--arch-body-col", default="V")
    parser.add_argument(
        "--preview-csv",
        action="store_true",
        help="Generate CSV previews before writing Excel files",
    )
    parser.add_argument(
        "--preview-only",
        action="store_true",
        help="Generate preview CSV files only, do not write Excel files",
    )

    args = parser.parse_args()

    base = Path(args.base_dir)
    req_md = base / "rd_03_010_requirements.md"
    arch_md = base / "rd_03_010_architecture_design.md"

    req_entries = parse_md(req_md)
    arch_entries = parse_md(arch_md)

    hsi_entries = filter_hsi(arch_entries)
    verification_entries = filter_verification(arch_entries)

    req_out = base / "rd_03_010_requirements_from_md.xlsx"
    arch_out = base / "rd_03_010_architecture_from_md.xlsx"
    hsi_out = base / "rd_03_010_hsi_from_md.xlsx"
    verification_out = base / "rd_03_010_verification_from_md.xlsx"

    req_preview_csv = base / "preview_requirements.csv"
    arch_preview_csv = base / "preview_architecture.csv"
    hsi_preview_csv = base / "preview_hsi.csv"
    verification_preview_csv = base / "preview_verification.csv"

    if args.preview_csv or args.preview_only:
        write_preview_csv(
            req_preview_csv,
            req_entries,
            title_col_letter=args.req_title_col,
            body_col_letter=args.req_body_col,
        )
        write_preview_csv(
            arch_preview_csv,
            arch_entries,
            title_col_letter=args.arch_title_col,
            body_col_letter=args.arch_body_col,
        )
        write_preview_csv(
            hsi_preview_csv,
            hsi_entries,
            title_col_letter=args.arch_title_col,
            body_col_letter=args.arch_body_col,
        )
        write_preview_csv(
            verification_preview_csv,
            verification_entries,
            title_col_letter=args.arch_title_col,
            body_col_letter=args.arch_body_col,
        )

        print(req_preview_csv)
        print(arch_preview_csv)
        print(hsi_preview_csv)
        print(verification_preview_csv)

    if args.preview_only:
        print("rows", len(req_entries), len(arch_entries), len(hsi_entries), len(verification_entries))
        return

    write_export_sheet(
        Path(args.req_template),
        req_out,
        req_entries,
        title_col_letter=args.req_title_col,
        body_col_letter=args.req_body_col,
    )

    write_export_sheet(
        Path(args.arch_template),
        arch_out,
        arch_entries,
        title_col_letter=args.arch_title_col,
        body_col_letter=args.arch_body_col,
    )

    write_export_sheet(
        Path(args.arch_template),
        hsi_out,
        hsi_entries,
        title_col_letter=args.arch_title_col,
        body_col_letter=args.arch_body_col,
    )

    write_export_sheet(
        Path(args.arch_template),
        verification_out,
        verification_entries,
        title_col_letter=args.arch_title_col,
        body_col_letter=args.arch_body_col,
    )

    print(req_out)
    print(arch_out)
    print(hsi_out)
    print(verification_out)
    print("rows", len(req_entries), len(arch_entries), len(hsi_entries), len(verification_entries))


if __name__ == "__main__":
    main()
